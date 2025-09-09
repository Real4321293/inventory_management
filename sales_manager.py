"""A sales manager file to handle
 the logic of product sold
 """

from openpyxl import load_workbook

class SalesManager:
    # Class  initialization
    def __init__(self, inventory_file = "inventory.xlsx"):
        self.inventory_file = inventory_file

    # Method to record the sales
    def record_sales(self, product_id, quantity_sold):
        try:
            wb = load_workbook(self.inventory_file)
            # Accessing the sheets

            inventory_sheet = wb['Inventory']
            sales_sheet = wb['Sales']

            # Flag to check if product is found
            product_found = False
            unit_price = None

            for row in inventory_sheet.iter_rows(min_row=2):
                if str(row[0].value) == str(product_id):
                    current_stock = row[3].value
                    unit_price = row[4].value

                    if quantity_sold <= current_stock:
                        # Update stock in the inventory
                        quantity_sold = int(quantity_sold)
                        current_stock = int(row[3].value)
                        new_stock = current_stock - quantity_sold
                        row[3].value = new_stock
                        product_found = True
                    else:
                        return "Error: Not Enough Stock"
                    break  # Stop looping after finding the product

            # If product was never found
            if not product_found:
                return "Error: The product is not found in the inventory"

            #  Generate Sale ID (outside the loop)
            sales_id = self.generate_sales_id(sales_sheet)

            #  Calculate total amount
            total_amount = quantity_sold * unit_price

            #  Insert into Sales sheet
            from datetime import date
            sales_sheet.append([sales_id, product_id, quantity_sold, date.today(), total_amount])

            last_row = sales_sheet.max_row
            sales_sheet.cell(row=last_row, column=5).number_format = '"Ksh" #,##0'

            #  Save workbook
            wb.save(self.inventory_file)



            return "Sale recorded successfully"
        except Exception as e:
            return f"Error recording sales: {str(e)}"




    # Method to auto generate sales id
    @staticmethod
    def generate_sales_id(ws_sales):
        last_row = ws_sales.max_row
        last_id = ws_sales.cell(row=last_row, column=1).value

        # If no sales yet or header only
        if not last_id or str(last_id).startswith("SaleID"):
            return "S001"

        try:
            num = int(str(last_id)[1:]) + 1
            return f"S{num:03d}"
        except:
            return "S001"

    # Method to get all the sales
    def get_all_sales(self):
        try:
            # loading  the workbook
            workbook = load_workbook(self.inventory_file)

            sales_sheet = workbook['Sales']

            if not sales_sheet:
                return "Error: No sales sheet found"

            sales_records = []

            # Reading all the rows and skipping the headers
            for row in sales_sheet.iter_rows(min_row=2, values_only=True):
                record = {
                    "SaleID": row[0],
                    "ProductID": row[1],
                    "QuantitySold": row[2],
                    "SaleDate": row[3],
                    "TotalAmount": row[4]
                }

                sales_records.append(record)

            return sales_records
        except Exception as e:
            return f"Error getting the  sales: {str(e)}"

# Creating an instance of the class sales manager
sales_manager = SalesManager()
sales_manager.record_sales(2, 10)
sales_manager.record_sales(3, 20)
sales_manager.get_all_sales()

