"""Class data manager to hold act as a database
for inventory and management system  excel is used ad the database"""

import os
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font


class DataManager:
    # Class initialization
    def __init__(self, inventory_file):
        self.inventory_file = inventory_file


    #Method initialize file
    def initialize_file(self):

        try:
            # Sheet names and sheet headers

            sheet_info = {
                "Inventory":["ProductID", "Name", "Category", "Quantity", "UnitPrice", "ReorderLevel"],
                "Sales":["SalesID", "ProductID", "QuantitySold", "SaleDate", "TotalAmount"]
            }

            # Checking if the file exists and creating if
            if not os.path.exists(self.inventory_file):
                wb = Workbook()

                # Removing the default sheet
                default_sheet = wb.active
                wb.remove(default_sheet)

                # Creating inventory and sales sheet with headers
                for sheet_name, headers in sheet_info.items():
                    ws = wb.create_sheet(sheet_name)
                    ws.append(headers)
                    self._format_headers(ws)

                wb.save(self.inventory_file)
                return "File created with inventory and sales sheet"
            else:
                # If file already exists
                wb = load_workbook(self.inventory_file)

                for sheet_name, headers in sheet_info.items():
                    if sheet_name not in wb.sheetnames:
                        # Creating a missing sheet with headers
                        ws = wb.create_sheet(sheet_name)
                        ws.append(headers)
                        self._format_headers(ws)
                    else:
                        # If headers  missing or wrong
                        ws = wb[sheet_name]

                        if list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0] != tuple(headers):
                            ws.delete_rows(1)
                            ws.append(headers)
                            self._format_headers(ws)

                wb.save(self.inventory_file)
                return "Sheets updated successfully"

        except PermissionError:
            return "Error: File is open in another program. Please close and try again."
        except Exception as e:
            return f"Error initializing file: {str(e)}"

    # Method to format headers
    @staticmethod
    def _format_headers(ws):
        # Applying bold font and adjusting headers to fit in the column widths
        for col, cell in enumerate(ws[1], start=1):# to specify first row
            cell.font = Font(bold=True)
            ws.column_dimensions[cell.column_letter].width = max(15, len(str(cell.value)))

    #  method to add product
    def add_product(self, product_id, name, category, quantity, unit_price, reorder_level):
        try:
            if not os.path.exists(self.inventory_file):
                return "Error: Inventory file does not exist"

            wb = load_workbook(self.inventory_file)# Loading the Workbook

            if "Inventory" not in wb.sheetnames:
                return "Error: Inventory sheet already exists"

            sheet = wb["Inventory"]# Loading the Worksheet

            # Check for duplicates product_id
            for cell in sheet.iter_rows(min_row=1, values_only=True):
                if cell[0] == product_id:
                    return f"Error : Product Id {product_id} already exists"

            # Appending the products
            sheet.append([product_id, name, category, quantity, unit_price, reorder_level])

            last_row = sheet.max_row
            sheet.cell(row=last_row, column=5).number_format = '"Ksh" #,##0'


            wb.save(self.inventory_file)
            return f"product {product_id} added to inventory"
        except Exception as e:
            return f"Error adding product: {str(e)}"

    # Method to update products
    def update_product(self, product_id, quantity= None, unit_price= None, reorder_level= None):
        try:
            if not os.path.exists(self.inventory_file):
                return "Error: Inventory file does not exist"

            wb = load_workbook(self.inventory_file)# Loading the work book

            if "Inventory" not in wb.sheetnames:
                return "Error: Inventory sheet does not exist"

            ws = wb["Inventory"] # Loading the work sheet

            product_found = False

            for row in ws.iter_rows(min_row=2, values_only=False):
                if row[0].value == product_id: # product column
                    product_found = True

                    if quantity is not None:
                        row[3].value = quantity # Quantity column

                    if unit_price is not None:
                        row[4].value = unit_price # Unit price column

                    if reorder_level is not None:
                        row[5].value = reorder_level

                        break


            if not product_found:
                return f"Error: {product_id} not found in inventory"

            wb.save(self.inventory_file)
            return f"product {product_id} updated successfully"

        except PermissionError:
            return "Error: Inventory sheet does not exist"
        except Exception as e:
            return f"Error updating product: {str(e)}"







manager = DataManager("inventory.xlsx")
manager.initialize_file()
manager.add_product(1, "Keyboard", "Computer Accessories", 30,  40000, "8 Weeks")
manager.add_product(2, "Mouse", "Computer Accessories", 30, 20000, "2 Weeks")
manager.add_product(3, "Monitor", "Computer Accessories", 40,  50000, "4 Weeks")
manager.update_product(1, quantity= 40)

